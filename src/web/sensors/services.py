import io
import math

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from constants import (
    PERCENTILE_LEVEL,
    TRIM_PERCENT,
    FORCE_THRESHOLD,
    KOEF_POWER,
    TEMPO_BORDER_PERCENT,
    DEGREE_POWER)

from database.models import Sprints


def build_sprint_hits_excel(
    slot_id: int, sprint_id: int, sprints: list[Sprints]
) -> bytes:
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = 'summary'

    devices_hits = {}
    for sp in sprints:
        device = sp.sensor_id or 'UNKNOWN'
        hits = []
        if isinstance(sp.data, dict):
            raw_hits = sp.data.get('hits', [])
            if isinstance(raw_hits, list):
                hits = raw_hits
        devices_hits[device] = hits

    summary_ws.append(['Slot ID', slot_id])
    summary_ws.append(['Sprint ID', sprint_id])
    summary_ws.append([])
    summary_ws.append(['Device', 'Hits Count'])
    total_all = 0
    for device, hits in devices_hits.items():
        summary_ws.append([device, len(hits)])
        total_all += len(hits)
    summary_ws.append([])
    summary_ws.append(['TOTAL', total_all])

    _auto_width(summary_ws)

    for device, hits in devices_hits.items():
        sheet_name = device[:31]
        ws = wb.create_sheet(title=sheet_name)

        all_keys = set()
        for h in hits:
            if isinstance(h, dict):
                all_keys.update(h.keys())

        if not all_keys:
            columns = ['raw']
        else:
            preferred = ['timeMs', 'maxAccel']
            rest = sorted(k for k in all_keys if k not in preferred)
            columns = [k for k in preferred if k in all_keys] + rest

        ws.append(['#', *columns])

        for idx, h in enumerate(hits, start=1):
            if isinstance(h, dict):
                row = [h.get(k) for k in columns]
            else:
                row = [str(h)]
            ws.append([idx, *row])

        _auto_width(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            v = cell.value
            if v is None:
                continue
            ln = len(str(v))
            if ln > max_len:
                max_len = ln
        ws.column_dimensions[letter].width = min(max_len + 2, 60)


def is_synced_hit(time_ms: int, blink_interval: float) -> bool:
    if time_ms is None or blink_interval <= 0:
        return False
    k = round(time_ms / blink_interval)
    nearest = k * blink_interval
    return abs(time_ms - nearest) <= blink_interval * TEMPO_BORDER_PERCENT


def get_forces_and_times(hits: list) -> tuple[list[float], list[int]]:
    forces_all, times_all = [], []
    for h in hits:
        f = h.get('maxAccel')
        if f is None:
            continue
        try:
            f = float(f)
        except (TypeError, ValueError):
            continue
        forces_all.append(f)
        t = h.get('timeMs')
        times_all.append(int(t) if t is not None else None)
    return forces_all, times_all


def get_filtered_forces(
    forces_all: list[float], force_threshold: float
) -> list[float]:
    return [f for f in forces_all if f >= force_threshold]


def _percentile(sorted_vals: list[float], q: float) -> float:
    n = len(sorted_vals)
    if n == 0:
        return 0.0
    if q <= 0:
        return float(sorted_vals[0])
    if q >= 100:
        return float(sorted_vals[-1])
    k = (n - 1) * (q / 100.0)
    f = math.floor(k)
    c = math.ceil(k)
    if f == c:
        return float(sorted_vals[int(k)])
    return sorted_vals[f] * (c - k) + sorted_vals[c] * (k - f)


def _trim_bounds(sorted_vals: list[float], TRIM_PERCENT: float):
    low_q  = TRIM_PERCENT * 100.0
    high_q = 100.0 - low_q
    return _percentile(sorted_vals, low_q), _percentile(sorted_vals, high_q)


def _trimmed(vals: list[float], trim_percent: float) -> list[float]:
    if not vals or trim_percent <= 0:
        return list(vals)
    s = sorted(vals)
    lo, hi = _trim_bounds(s, trim_percent)
    return [x for x in s if lo <= x <= hi]

def calculate_sprint_metrics(
    hits: list,
    blink_interval: float,
    hit_count: int,
    force_threshold: float | None = None,
    trim_percent: float | None = None,
    percentile_level: float | None = None,
) -> dict:
    force_threshold = force_threshold or FORCE_THRESHOLD
    trim_percent = trim_percent or TRIM_PERCENT
    percentile_level = percentile_level or PERCENTILE_LEVEL
    if not hits or hit_count == 0:
        return {}

    forces_all, times_all = get_forces_and_times(hits)
    if not forces_all:
        return {}

    forces = get_filtered_forces(forces_all, force_threshold)
    # if not forces:
    #     return {'tempo': 0.0, 'power': 0.0, 'energy': 0.0}

    trimmed_for_Fref = _trimmed(forces, trim_percent) or list(forces)
    F_ref = _percentile(sorted(trimmed_for_Fref), percentile_level)

    if F_ref > 0:
        mean_square_ratio = sum((f / F_ref) ** 2 for f in forces) / len(forces)
        EI_norm_ref = 100.0 * mean_square_ratio
    else:
        EI_norm_ref = 0.0

    power = EI_norm_ref

    max_punch = max(forces_all)
    sum_punches = sum(forces_all)
    average_punch = sum_punches / hit_count if hit_count > 0 else 0
    power_old = (average_punch / max_punch) * KOEF_POWER if max_punch > 0 else 0

    base = [(t, f) for t, f in zip(times_all, forces_all) if t is not None]
    if base:
        synced = sum(1 for (t, _) in base if is_synced_hit(int(t), blink_interval))
        tempo = (synced / len(base)) * 100.0
    else:
        tempo = 0.0
    energy_old = tempo * (power_old / KOEF_POWER) ** DEGREE_POWER
    energy = tempo * (power / KOEF_POWER) ** DEGREE_POWER
    return {
        'tempo': round(tempo, 2),
        'power': round(power, 2),
        'energy': round(energy, 2),
        'power_old': round(power_old, 2),
        'energy_old': round(energy_old, 2),
    }
